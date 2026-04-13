from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any
import xml.etree.ElementTree as ET
import zipfile

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None


DATA_DIR = Path(__file__).resolve().parents[1] / "data"
DEFAULT_WORKBOOK_PATH = DATA_DIR / "test_data.xlsx"
PRODUCT_SHEET_ALIASES = ("products", "product", "favourites")


@dataclass(frozen=True)
class LoginData:
    username: str
    password: str
    expected_url_after_login: str


@dataclass(frozen=True)
class ProductData:
    sku: str
    product_name: str
    expected_price: str

    @property
    def price_value(self) -> float:
        cleaned_value = self.expected_price.replace("$", "").replace(",", "").strip()
        return round(float(cleaned_value), 2)


class ExcelDataReader:
    def __init__(self, workbook_path: str | Path | None = None):
        self.workbook_path = Path(workbook_path or DEFAULT_WORKBOOK_PATH)
        self._sheet_cache: dict[str, list[dict[str, Any]]] | None = None

    def get_login_data(self) -> LoginData:
        rows = self.get_sheet_rows("login")
        if not rows:
            raise ValueError("The 'login' sheet does not contain any rows.")
        row = rows[0]
        return LoginData(
            username=str(row["username"]),
            password=str(row["password"]),
            expected_url_after_login=str(row["expected_url_after_login"]),
        )

    def get_products(self, names: list[str] | None = None, limit: int | None = None) -> list[ProductData]:
        rows = self.get_sheet_rows("products", aliases=PRODUCT_SHEET_ALIASES)
        products = [
            ProductData(
                sku=str(row["sku"]),
                product_name=str(row["product_name"]),
                expected_price=str(row["expected_price"]),
            )
            for row in rows
        ]

        if names:
            name_set = {name.casefold() for name in names}
            products = [product for product in products if product.product_name.casefold() in name_set]

        if limit is not None:
            products = products[:limit]

        return products

    def get_sheet_rows(self, sheet_name: str, aliases: tuple[str, ...] = ()) -> list[dict[str, Any]]:
        normalized_data = self._load_workbook_data()
        for candidate in (sheet_name, *aliases):
            rows = normalized_data.get(candidate.casefold())
            if rows is not None:
                return rows

        available = ", ".join(sorted(normalized_data))
        raise ValueError(
            f"Could not find sheet '{sheet_name}' in '{self.workbook_path.name}'. Available sheets: {available}"
        )

    def _load_workbook_data(self) -> dict[str, list[dict[str, Any]]]:
        if self._sheet_cache is None:
            if not self.workbook_path.exists():
                raise FileNotFoundError(f"Workbook not found: {self.workbook_path}")

            if load_workbook is not None:
                self._sheet_cache = self._load_with_openpyxl()
            else:
                self._sheet_cache = self._load_from_xlsx_archive()

        return self._sheet_cache

    def _load_with_openpyxl(self) -> dict[str, list[dict[str, Any]]]:
        workbook = load_workbook(self.workbook_path, read_only=True, data_only=True)
        parsed: dict[str, list[dict[str, Any]]] = {}
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            rows = list(worksheet.iter_rows(values_only=True))
            parsed[sheet_name.casefold()] = self._rows_to_dicts(rows)
        workbook.close()
        return parsed

    def _load_from_xlsx_archive(self) -> dict[str, list[dict[str, Any]]]:
        namespace = {
            "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
            "rel": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
            "pkg": "http://schemas.openxmlformats.org/package/2006/relationships",
        }
        parsed: dict[str, list[dict[str, Any]]] = {}

        with zipfile.ZipFile(self.workbook_path) as workbook_archive:
            shared_strings = self._load_shared_strings(workbook_archive, namespace)
            workbook_root = ET.fromstring(workbook_archive.read("xl/workbook.xml"))
            rels_root = ET.fromstring(workbook_archive.read("xl/_rels/workbook.xml.rels"))
            relation_map = {
                relation.attrib["Id"]: relation.attrib["Target"]
                for relation in rels_root.findall("pkg:Relationship", namespace)
            }

            for sheet in workbook_root.findall("main:sheets/main:sheet", namespace):
                sheet_name = sheet.attrib["name"]
                relation_id = sheet.attrib["{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"]
                sheet_path = "xl/" + relation_map[relation_id]
                sheet_root = ET.fromstring(workbook_archive.read(sheet_path))
                rows = self._extract_rows(sheet_root, shared_strings, namespace)
                parsed[sheet_name.casefold()] = self._rows_to_dicts(rows)

        return parsed

    @staticmethod
    def _load_shared_strings(workbook_archive: zipfile.ZipFile, namespace: dict[str, str]) -> list[str]:
        if "xl/sharedStrings.xml" not in workbook_archive.namelist():
            return []

        root = ET.fromstring(workbook_archive.read("xl/sharedStrings.xml"))
        values: list[str] = []
        for item in root.findall("main:si", namespace):
            parts = [text_node.text or "" for text_node in item.findall(".//main:t", namespace)]
            values.append("".join(parts))
        return values

    @staticmethod
    def _extract_rows(
        sheet_root: ET.Element,
        shared_strings: list[str],
        namespace: dict[str, str],
    ) -> list[list[Any]]:
        rows: list[list[Any]] = []
        for row in sheet_root.findall("main:sheetData/main:row", namespace):
            values: list[Any] = []
            for cell in row.findall("main:c", namespace):
                cell_type = cell.attrib.get("t")
                value_node = cell.find("main:v", namespace)
                value = None if value_node is None else value_node.text
                if cell_type == "s" and value is not None:
                    value = shared_strings[int(value)]
                values.append(value)
            rows.append(values)
        return rows

    @staticmethod
    def _rows_to_dicts(rows: list[Any]) -> list[dict[str, Any]]:
        if not rows:
            return []

        headers = [str(value).strip() for value in rows[0] if value is not None]
        results: list[dict[str, Any]] = []

        for row in rows[1:]:
            values = list(row)
            if not any(value not in (None, "") for value in values):
                continue

            row_dict: dict[str, Any] = {}
            for index, header in enumerate(headers):
                row_dict[header] = "" if index >= len(values) or values[index] is None else values[index]
            results.append(row_dict)

        return results
